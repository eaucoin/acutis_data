import os
import sys
import json
import argparse
from bs4 import BeautifulSoup
import pandas as pd
import openpyxl
import xlrd
import re
from striprtf.striprtf import rtf_to_text
import mammoth
from dataclasses import dataclass
from typing import List, Dict, Any, Tuple, Set, Optional
import extract_msg
from email.utils import parseaddr
from pathlib import Path
from tqdm import tqdm

# Excel to JSON Processing Classes and Functions
@dataclass
class CellCoordinate:
    row: int
    col: int

    def __hash__(self):
        return hash((self.row, self.col))

class ExcelProcessor:
    def __init__(self, file_path: str):
        self.file_path = file_path
        self.file_extension = os.path.splitext(file_path)[1].lower()

    def is_cell_empty(self, value: Any) -> bool:
        if value is None:
            return True
        if isinstance(value, str) and value.strip() == "":
            return True
        return False

    # [Previous Excel Processing Methods - Same as before]
    def get_continuous_bodies(self, data: List[List[Any]]) -> List[Set[CellCoordinate]]:
        """Find continuous bodies of non-empty cells using flood fill algorithm."""
        def is_valid(r: int, c: int) -> bool:
            return 0 <= r < len(data) and 0 <= c < len(data[0])

        def get_neighbors(coord: CellCoordinate) -> List[CellCoordinate]:
            r, c = coord.row, coord.col
            directions = [(0, 1), (1, 0), (0, -1), (-1, 0)]  # right, down, left, up
            return [CellCoordinate(r + dr, c + dc) for dr, dc in directions 
                   if is_valid(r + dr, c + dc) and not self.is_cell_empty(data[r + dr][c + dc])]

        def flood_fill(start: CellCoordinate, visited: Set[CellCoordinate]) -> Set[CellCoordinate]:
            if start in visited or self.is_cell_empty(data[start.row][start.col]):
                return set()
            
            body = {start}
            to_visit = [start]
            visited.add(start)

            while to_visit:
                current = to_visit.pop(0)
                for neighbor in get_neighbors(current):
                    if neighbor not in visited:
                        visited.add(neighbor)
                        body.add(neighbor)
                        to_visit.append(neighbor)
            
            return body

        visited = set()
        bodies = []
        
        for r in range(len(data)):
            for c in range(len(data[0])):
                coord = CellCoordinate(r, c)
                if coord not in visited and not self.is_cell_empty(data[r][c]):
                    body = flood_fill(coord, visited)
                    if body:
                        bodies.append(body)
        
        return bodies

    def find_left_cells(self, bodies: List[Set[CellCoordinate]]) -> Set[CellCoordinate]:
        """Find leftmost cells for each row in continuous bodies."""
        left_cells = set()
        for body in bodies:
            rows = {}
            for coord in body:
                if coord.row not in rows or coord.col < rows[coord.row].col:
                    rows[coord.row] = coord
            left_cells.update(rows.values())
        return left_cells

    def find_top_cells(self, bodies: List[Set[CellCoordinate]]) -> Set[CellCoordinate]:
        """Find topmost cells for each column in continuous bodies."""
        top_cells = set()
        for body in bodies:
            cols = {}
            for coord in body:
                if coord.col not in cols or coord.row < cols[coord.col].row:
                    cols[coord.col] = coord
            top_cells.update(cols.values())
        return top_cells

    def format_label_value(self, value: Any) -> str:
        """Format a value for use as a label."""
        if isinstance(value, (int, float)):
            return str(value)
        return f'"{str(value)}"'
    
    def find_unique_leftmost_cells(self, continuous_body: Set[CellCoordinate]) -> List[CellCoordinate]:
        """Find cells that are uniquely in their column and are leftmost in the body."""
        # Group cells by column
        column_groups = {}
        for cell in continuous_body:
            if cell.col not in column_groups:
                column_groups[cell.col] = []
            column_groups[cell.col].append(cell)
        
        # Find leftmost columns that have only one cell
        result = []
        min_col = min(cell.col for cell in continuous_body)
        for col, cells in column_groups.items():
            # Check if this column:
            # 1. Has exactly one cell
            # 2. Is to the left of at least one other column in the body
            if len(cells) == 1 and any(other_col > col for other_col in column_groups.keys()):
                result.append(cells[0])
        
        return result

    def find_unique_topmost_cells(self, continuous_body: Set[CellCoordinate]) -> List[CellCoordinate]:
        """Find cells that are uniquely in their row and are topmost in the body."""
        # Group cells by row
        row_groups = {}
        for cell in continuous_body:
            if cell.row not in row_groups:
                row_groups[cell.row] = []
            row_groups[cell.row].append(cell)
        
        # Find topmost rows that have only one cell
        result = []
        min_row = min(cell.row for cell in continuous_body)
        for row, cells in row_groups.items():
            # Check if this row:
            # 1. Has exactly one cell
            # 2. Is above at least one other row in the body
            if len(cells) == 1 and any(other_row > row for other_row in row_groups.keys()):
                result.append(cells[0])
        
        return result

    def get_cell_components(self, coord: CellCoordinate, data: List[List[Any]], 
                        left_cells: Set[CellCoordinate], top_cells: Set[CellCoordinate],
                        current_body: Set[CellCoordinate]) -> Dict[str, Any]:
        """Generate three-component information for a cell."""
        x_labels = []
        y_labels = []
        
        # First, get unique special cells
        unique_leftmost = self.find_unique_leftmost_cells(current_body)
        unique_topmost = self.find_unique_topmost_cells(current_body)
        
        # Add unique leftmost values first (if any exist)
        if unique_leftmost:
            x_labels.extend([
                self.format_label_value(data[cell.row][cell.col])
                for cell in sorted(unique_leftmost, key=lambda c: abs(c.row - coord.row))
            ])
        
        # Add unique topmost values first (if any exist)
        if unique_topmost:
            y_labels.extend([
                self.format_label_value(data[cell.row][cell.col])
                for cell in sorted(unique_topmost, key=lambda c: abs(c.col - coord.col))
            ])
        
        # Add regular left cells (in order of proximity)
        regular_left_cells = sorted(
            [c for c in left_cells if c.row == coord.row and c.col < coord.col],
            key=lambda c: coord.col - c.col
        )
        if regular_left_cells:
            x_labels.extend([
                self.format_label_value(data[c.row][c.col])
                for c in regular_left_cells
            ])
        
        # Add regular top cells (in order of proximity)
        regular_top_cells = sorted(
            [c for c in top_cells if c.col == coord.col and c.row < coord.row],
            key=lambda c: coord.row - c.row
        )
        if regular_top_cells:
            y_labels.extend([
                self.format_label_value(data[c.row][c.col])
                for c in regular_top_cells
            ])
        
        # Cell qualifies if it has either unique cells or regular label cells
        if x_labels or y_labels or unique_leftmost or unique_topmost:
            return {
                "x_labels": " OR ".join(x_labels) if x_labels else "None",
                "y_labels": " OR ".join(y_labels) if y_labels else "None",
                "value": data[coord.row][coord.col]
            }
        return None

    def process_sheet(self, data: List[List[Any]]) -> List[Dict[str, Any]]:
        """Process a single sheet and return its cell data."""
        # Get continuous bodies
        bodies = self.get_continuous_bodies(data)
        
        # Create a map of cell coordinates to their continuous body
        cell_to_body = {}
        for body in bodies:
            for cell in body:
                cell_to_body[cell] = body
        
        # Get left and top cells as before
        left_cells = self.find_left_cells(bodies)
        top_cells = self.find_top_cells(bodies)
        
        cell_data = []
        for r in range(len(data)):
            for c in range(len(data[0])):
                if not self.is_cell_empty(data[r][c]):
                    coord = CellCoordinate(r, c)
                    if coord in cell_to_body:  # If cell is part of any continuous body
                        components = self.get_cell_components(
                            coord, data, left_cells, top_cells, cell_to_body[coord]
                        )
                        if components:
                            cell_data.append(components)
        
        return cell_data

    def convert_to_json(self) -> Dict[str, List[Dict[str, Any]]]:
        """Convert Excel file to JSON format."""
        result = {}
        
        try:
            if self.file_extension == '.xls':
                workbook = xlrd.open_workbook(self.file_path)
                for sheet_name in workbook.sheet_names():
                    worksheet = workbook.sheet_by_name(sheet_name)
                    data = [[worksheet.cell(r, c).value 
                            for c in range(worksheet.ncols)]
                           for r in range(worksheet.nrows)]
                    result[sheet_name] = self.process_sheet(data)
                    
            else:  # xlsx files
                workbook = openpyxl.load_workbook(self.file_path)
                for sheet_name in workbook.sheetnames:
                    worksheet = workbook[sheet_name]
                    data = [[cell.value for cell in row] 
                           for row in worksheet.rows]
                    result[sheet_name] = self.process_sheet(data)
                    
        except Exception as e:
            print(f"Error processing {self.file_path}: {str(e)}")
            return None
            
        return result

# Original Conversion Functions for Other Formats
def html_to_markdown(html_content):
    soup = BeautifulSoup(html_content, 'html.parser')
    markdown_content = ""

    if not is_program_generated_html(soup):
        return None

    for page in soup.find_all('div', class_='page'):
        for element in page.children:
            if element.name == 'div' and 'page-number' in element.get('class', []):
                markdown_content += convert_page_number(element)
            elif element.name == 'h1' and 'title' in element.get('class', []):
                markdown_content += convert_title(element)
            elif element.name == 'h2' and 'section-header' in element.get('class', []):
                markdown_content += convert_section_header(element)
            elif element.name == 'p' and 'text' in element.get('class', []):
                markdown_content += convert_paragraph(element)
            elif element.name == 'li' and 'list-item' in element.get('class', []):
                markdown_content += convert_list_item(element)
            elif element.name == 'figcaption' and 'caption' in element.get('class', []):
                markdown_content += convert_caption(element)
            elif element.name == 'div' and 'footnote' in element.get('class', []):
                markdown_content += convert_footnote(element)
            elif element.name == 'div' and 'formula' in element.get('class', []):
                markdown_content += convert_formula(element)
            elif element.name == 'footer' and 'page-footer' in element.get('class', []):
                markdown_content += convert_footer(element)
            elif element.name == 'header' and 'page-header' in element.get('class', []):
                markdown_content += convert_header(element)
            elif element.name == 'table' and 'data-table' in element.get('class', []):
                markdown_content += convert_table(element)

        if page != soup.find_all('div', class_='page')[-1]:
            markdown_content += "\n---\n\n"

    return markdown_content

def is_program_generated_html(soup):
    if soup.find_all('div', class_='page') and soup.find('div', class_='page-number'):
        return True
    return False

def convert_title(element):
    return f"# {element.get_text().strip()}\n\n"

def convert_section_header(element):
    return f"## {element.get_text().strip()}\n\n"

def convert_paragraph(element):
    return f"{element.get_text().strip()}\n\n"

def convert_list_item(element):
    return f"- {element.get_text().strip()}\n"

def convert_caption(element):
    return f"*{element.get_text().strip()}*\n\n"

def convert_footnote(element):
    return f"[Footnote: {element.get_text().strip()}]\n\n"

def convert_formula(element):
    return f"${element.get_text().strip()}$\n\n"

def convert_footer(element):
    return f"---\n*{element.get_text().strip()}*\n\n"

def convert_header(element):
    return f"*{element.get_text().strip()}*\n---\n\n"

def convert_page_number(element):
    return f"<!-- {element.get_text().strip()} -->\n\n"

def convert_table(table):
    markdown_table = ""
    rows = table.find_all('tr')
    if not rows:
        return ""

    max_cells = max(len(row.find_all('td')) for row in rows)
    markdown_table += "| " + " | ".join([""] * max_cells) + " |\n"
    markdown_table += "| " + " | ".join(["---"] * max_cells) + " |\n"

    for row in rows:
        cells = row.find_all('td')
        markdown_row = "| " + " | ".join(cell.get_text().strip() for cell in cells)
        markdown_row += " | " * (max_cells - len(cells)) + " |\n"
        markdown_table += markdown_row

    return markdown_table + "\n"

def convert_rtf_to_markdown(rtf_path):
    try:
        with open(rtf_path, 'r', encoding='utf-8', errors='ignore') as file:
            rtf_text = file.read()
        
        plain_text = rtf_to_text(rtf_text)
        lines = plain_text.split('\n')
        markdown_lines = []
        i = 0
        while i < len(lines):
            line = lines[i].strip()
            if line:
                if i + 1 >= len(lines) or not lines[i + 1].strip():
                    if len(line) < 50:
                        if line.isupper():
                            markdown_lines.append(f"# {line.title()}\n")
                        else:
                            markdown_lines.append(f"## {line}\n")
                    else:
                        markdown_lines.append(line + "\n")
                else:
                    markdown_lines.append(line + "\n")
            else:
                markdown_lines.append("\n")
            i += 1
        
        markdown_content = ""
        in_list = False
        for line in markdown_lines:
            stripped_line = line.strip()
            if stripped_line.startswith(('•', '-', '*', '○', '·', '>', '▪')):
                if not in_list:
                    markdown_content += "\n"
                    in_list = True
                markdown_content += f"- {stripped_line[1:].strip()}\n"
            elif stripped_line.startswith(tuple(f"{i}." for i in range(10))):
                if not in_list:
                    markdown_content += "\n"
                    in_list = True
                number, text = stripped_line.split('.', 1)
                markdown_content += f"{number}. {text.strip()}\n"
            else:
                if in_list and stripped_line:
                    markdown_content += "\n"
                in_list = False
                markdown_content += line
        
        return markdown_content
        
    except Exception as e:
        print(f"Error converting RTF file {rtf_path}: {str(e)}")
        return None

def convert_docx_to_markdown(docx_path):
    try:
        style_map = """
        p[style-name='heading 1'] => h1:fresh
        p[style-name='heading 2'] => h2:fresh
        p[style-name='heading 3'] => h3:fresh
        p[style-name='heading 4'] => h4:fresh
        p[style-name='heading 5'] => h5:fresh
        p[style-name='heading 6'] => h6:fresh
        p[style-name='Title'] => h1:fresh
        p[style-name='Subtitle'] => h2:fresh
        p[style-name='List Bullet'] => ul > li:fresh
        p[style-name='List Number'] => ol > li:fresh
        p[style-name='Quote'] => blockquote:fresh
        p[style-name='Intense Quote'] => blockquote.intense:fresh
        p[style-name='Normal'] => p:fresh
        b => strong
        i => em
        strike => s
        u => u
        """
        
        with open(docx_path, "rb") as docx_file:
            result = mammoth.convert_to_markdown(
                docx_file,
                style_map=style_map,
                ignore_empty_paragraphs=True
            )
            
            important_messages = [
                msg for msg in result.messages 
                if "style" not in msg.message.lower() 
                and "message" in msg.message.lower()
            ]
            
            if important_messages:
                print(f"Important messages for {docx_path}:")
                for message in important_messages:
                    print(f"- {message}")
            
            markdown_content = result.value
            markdown_content = re.sub(r'\n{3,}', '\n\n', markdown_content)
            markdown_content = re.sub(r'(\n#{1,6}\s.*)\n(?=\S)', r'\1\n\n', markdown_content)
            markdown_content = re.sub(r'(\n[-*]\s.*)\n(?=[-*]\s)', r'\1\n', markdown_content)
            
            return markdown_content.strip() + '\n'
            
    except Exception as e:
        print(f"Error converting DOCX file {docx_path}: {str(e)}")
        return None

def convert_msg_to_markdown(msg_path):
    try:
        msg = extract_msg.Message(msg_path)
        markdown_parts = []
        
        markdown_parts.append("# " + (msg.subject or "No Subject"))
        markdown_parts.append("\n## Email Details\n")
        
        markdown_parts.append("| Field | Value |")
        markdown_parts.append("|-------|-------|")
        
        sender_name, sender_email = parseaddr(msg.sender)
        from_value = f"{sender_name} <{sender_email}>" if sender_name else sender_email
        markdown_parts.append(f"| From | {from_value} |")
        
        if msg.to:
            to_addresses = msg.to
            if isinstance(to_addresses, str):
                to_addresses = [to_addresses]
            markdown_parts.append(f"| To | {'; '.join(to_addresses)} |")
        
        if msg.cc:
            cc_addresses = msg.cc
            if isinstance(cc_addresses, str):
                cc_addresses = [cc_addresses]
            markdown_parts.append(f"| CC | {'; '.join(cc_addresses)} |")
        
        if msg.date:
            try:
                date_str = msg.date.strftime("%Y-%m-%d %H:%M:%S")
                markdown_parts.append(f"| Date | {date_str} |")
            except:
                pass
        
        markdown_parts.append("")
        
        if msg.body:
            markdown_parts.append("\n## Message Body\n")
            body_text = msg.body
            body_lines = body_text.split('\n')
            processed_lines = []
            
            for line in body_lines:
                if line.strip().startswith('>'):
                    processed_lines.append(line)
                elif line.strip().startswith('----'):
                    processed_lines.append('\n---\n')
                else:
                    processed_lines.append(line)
            
            body_markdown = '\n'.join(processed_lines)
            body_markdown = re.sub(r'\n{3,}', '\n\n', body_markdown)
            markdown_parts.append(body_markdown)
        
        if msg.attachments:
            markdown_parts.append("\n## Attachments\n")
            for attachment in msg.attachments:
                filename = attachment.longFilename or attachment.shortFilename
                if filename:
                    markdown_parts.append(f"- 📎 `{filename}`")
            markdown_parts.append("")
        
        msg.close()
        
        markdown_content = '\n'.join(markdown_parts)
        markdown_content = markdown_content.replace('\r', '')
        markdown_content = re.sub(r'\n{3,}', '\n\n', markdown_content)
        
        return markdown_content.strip() + '\n'
        
    except Exception as e:
        print(f"Error converting MSG file {msg_path}: {str(e)}")
        return None

def process_directory(input_dir: str) -> None:
    """Process all supported files in a directory."""
    files_to_process = []
    for root, _, files in os.walk(input_dir):
        for file in files:
            file_path = os.path.join(root, file)
            
            # Excel files (JSON output)
            if file.lower().endswith(('.xlsx', '.xls')):
                json_path = os.path.splitext(file_path)[0] + '.json'
                if os.path.exists(json_path):
                    print(f"Skipping {file_path} as JSON file already exists")
                    continue
                files_to_process.append((file_path, 'excel'))
                
            # Other formats (Markdown output)
            elif file.lower().endswith('.html'):
                md_path = os.path.splitext(file_path)[0] + '.md'
                if os.path.exists(md_path):
                    continue
                files_to_process.append((file_path, 'html'))
            elif file.lower().endswith('.rtf'):
                files_to_process.append((file_path, 'rtf'))
            elif file.lower().endswith('.docx') and not file.startswith('~$'):
                files_to_process.append((file_path, 'docx'))
            elif file.lower().endswith('.msg'):
                files_to_process.append((file_path, 'msg'))

    if not files_to_process:
        print("No files found to process")
        return

    print(f"Found {len(files_to_process)} files to process")
    
    successful = 0
    failed = 0
    
    for file_path, file_type in tqdm(files_to_process, desc="Converting files"):
        try:
            if file_type == 'excel':
                processor = ExcelProcessor(file_path)
                output_data = processor.convert_to_json()
                if output_data:
                    output_path = os.path.splitext(file_path)[0] + '.json'
                    with open(output_path, 'w', encoding='utf-8') as f:
                        json.dump(output_data, f, indent=2, ensure_ascii=False)
                    successful += 1
                else:
                    failed += 1
            else:
                # Handle other formats (original markdown conversion)
                md_path = os.path.splitext(file_path)[0] + '.md'
                markdown_content = None
                
                if file_type == 'html':
                    with open(file_path, 'r', encoding='utf-8') as f:
                        html_content = f.read()
                    markdown_content = html_to_markdown(html_content)
                elif file_type == 'rtf':
                    markdown_content = convert_rtf_to_markdown(file_path)
                elif file_type == 'docx':
                    markdown_content = convert_docx_to_markdown(file_path)
                elif file_type == 'msg':
                    markdown_content = convert_msg_to_markdown(file_path)

                if markdown_content:
                    os.makedirs(os.path.dirname(os.path.abspath(md_path)), exist_ok=True)
                    with open(md_path, 'w', encoding='utf-8') as f:
                        f.write(markdown_content)
                    successful += 1
                else:
                    failed += 1
                    
        except Exception as e:
            print(f"Error processing {file_path}: {str(e)}")
            failed += 1
    
    print(f"\nConversion complete: {successful} successful, {failed} failed")

def main():
    parser = argparse.ArgumentParser(
        description="Convert various document formats to Markdown or JSON"
    )
    parser.add_argument("input_dir", help="Path to the input directory containing files to convert")
    args = parser.parse_args()

    if not os.path.isdir(args.input_dir):
        print(f"Error: {args.input_dir} is not a valid directory.")
        sys.exit(1)

    process_directory(args.input_dir)

if __name__ == "__main__":
    main()
